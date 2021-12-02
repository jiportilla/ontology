#!/usr/bin/env python
# -*- coding: utf-8 -*-


from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Color
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.workbook.workbook import Workbook

from base import FileIO


class FormatGenerator(object):
    """ defines excel workbook formats to be applied in a generated feedback """

    def __init__(self,
                 some_workbook: Workbook,
                 some_style_config: str):
        """
        :param some_workbook:
            the excel workbook constructed in a prior step
        """
        if not some_workbook:
            raise ValueError("Mandatory Param: Workbook")
        if not some_style_config:
            raise ValueError("Style Configuration")

        self.workbook = some_workbook
        self.config = FileIO.file_to_yaml_by_relative_path(
            some_style_config)

    def process(self):
        """ define the formats to use in the feedback """

        def _side(d_side: dict) -> Side:
            return Side(style=d_side["style"],
                        color=d_side["color"])

        def _border(a_side: Side) -> Border:
            return Border(left=a_side,
                          top=a_side,
                          right=a_side,
                          bottom=a_side)

        def _fill(d_fill: dict) -> PatternFill:
            color = Color(rgb=d_fill['color'])
            return PatternFill(patternType='solid',
                               fgColor=color)

        for k in self.config:
            _named_style = NamedStyle(name=k)
            # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.alignment.html

            if "align" in self.config[k]:
                _named_style.alignment = AlignmentGenerator(self.config[k]["align"]).process()

            if "fill" in self.config[k]:
                _named_style.fill = _fill(self.config[k]["fill"])

            _named_style.font = FontGenerator(self.config[k]["font"]).process()
            if "side" in self.config[k]:
                _named_style.border = _border(_side(self.config[k]["side"]))

            if "number_format" in self.config[k]:
                _named_style.number_format = self.config[k]["number_format"]

            self.workbook.add_named_style(_named_style)


class AlignmentGenerator(object):
    horizontal_values = ['distributed',
                         'justify',
                         'center',
                         'left',
                         'fill',
                         'centerContinuous',
                         'right',
                         'general']

    vertical_values = ['bottom',
                       'distributed',
                       'justify',
                       'center',
                       'top']

    default_horizontal_style = "general"
    default_vertical_style = "bottom"
    default_text_rotation = 0
    default_shrink_to_fit = False
    default_indent = 0.0
    default_wrap_text = False

    def __init__(self,
                 d_align: dict):
        self.d_align = d_align

    def _horizontal(self) -> str:
        if "horizontal" in self.d_align:
            if self.d_align["horizontal"] not in self.horizontal_values:
                raise ValueError("\n".join([
                    "Unexpected Horizontal Value",
                    "\tactual: {}".format(self.d_align["horizontal"])
                ]))

            return self.d_align["horizontal"]
        return self.default_horizontal_style

    def _vertical(self) -> str:
        if "vertical" in self.d_align:
            if self.d_align["vertical"] not in self.vertical_values:
                raise ValueError("\n".join([
                    "Unexpected Vertical Value",
                    "\tactual: {}".format(self.d_align["vertical"])
                ]))

            return self.d_align["vertical"]
        return self.default_vertical_style

    def _text_rotation(self) -> int:
        if "textRotation" in self.d_align:
            return self.d_align["textRotation"]
        return self.default_text_rotation

    def _shrink_to_fit(self) -> bool:
        if "shrinkToFit" in self.d_align:
            return self.d_align["shrinkToFit"]
        return self.default_shrink_to_fit

    def _indent(self) -> float:
        if "indent" in self.d_align:
            return self.d_align["indent"]
        return self.default_shrink_to_fit

    def _wrap_text(self) -> float:
        if "wrap_text" in self.d_align:
            return self.d_align["wrap_text"]
        return self.default_wrap_text

    def process(self) -> Alignment:
        # https://openpyxl.readthedocs.io/en/stable/styles.html
        return Alignment(indent=self._indent(),
                         vertical=self._vertical(),
                         wrap_text=self._wrap_text(),
                         horizontal=self._horizontal(),
                         shrinkToFit=self._shrink_to_fit(),
                         textRotation=self._text_rotation())


class FontGenerator(object):
    default_font_name = "Arial"
    default_font_size = 12
    default_font_color = "000000"
    default_font_bold = False
    default_font_italics = False

    def __init__(self,
                 d_font: dict):
        self.d_font = d_font

    def _name(self) -> str:
        if "name" in self.d_font:
            return self.d_font["name"]
        return self.default_font_name

    def _bold(self) -> bool:
        if "bold" in self.d_font:
            return self.d_font["bold"]
        return self.default_font_bold

    def _italic(self) -> bool:
        if "italic" in self.d_font:
            return self.d_font["italic"]
        return self.default_font_italics

    def _size(self) -> int:
        if "size" in self.d_font:
            return self.d_font["size"]
        return self.default_font_size

    def _color(self) -> str:
        if "color" in self.d_font:
            return self.d_font["color"]
        return self.default_font_color

    def process(self) -> Font:
        return Font(name=self._name(),
                    bold=self._bold(),
                    italic=self._italic(),
                    size=self._size(),
                    color=self._color())
